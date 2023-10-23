import openpyxl
import warnings
import datetime
import numpy as np
import pandas as pd
from admm import run_admm
from data_preprocess import data_output
from data_preprocess import data_to_excel
from data_preprocess import data_preprocess
from data_preprocess import data_clustering
from data_preprocess import abundance_exchange
from data_estimation import diff_admm
from data_estimation import perf_admm
from data_plot import plot_loss_function
from data_plot import plot_jsd_iteration
from data_plot import plot_diff_xwh_iteration
from data_plot import plot_w_origin_iteration
from data_plot import plot_unknown_iteration
warnings.filterwarnings('ignore')


# setting up an excel datasheet for data padding
workbook_project_data = openpyxl.load_workbook('./project_data.xlsx')
workbook_feast_data = openpyxl.load_workbook('./feast_data.xlsx')
workbook_noise_data = openpyxl.load_workbook('./noise_data.xlsx')
workbook_sink_abundance = openpyxl.load_workbook('./sink_abundance.xlsx')
workbook_remove_data = openpyxl.load_workbook('./remove_data.xlsx')
workbook_w = openpyxl.load_workbook('./w.xlsx')
workbook_h = openpyxl.load_workbook('./h.xlsx')
workbook_w_plus = openpyxl.load_workbook('./w_plus.xlsx')
workbook_abundance = openpyxl.load_workbook('./estimated_abundance.xlsx')


# initialization parameter setting
data_path = './0.65jsd.xlsx'  # Datasets for source data, unknown data and sink data
proportion_path = './true_proportions.xlsx'
mode = 'normal'  # Data preprocessing using clustering
sink_number = 9  # The number of sink data
iteration = 4000  # Number of iterations of data per round
jsd_value = 0.25  # Setting the threshold for clustering
value = 100000  # Number of multinomial distribution samples for generating noise data
rho = 1
threshold = 0.000001  # threshold of convergence of Lagrangian functions
unknown_proportion = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]  # unknown proportion setting


# data output and storing: project_data, remove_data, feast_data, sink_abundance
if mode == 'cluster':
    project_data, feast_data, noise_sink_data, source_jsd, sources_num, noise_source_jsd = data_output(data_path,
                                                                                                       sink_number,
                                                                                                       proportion_path,
                                                                                                       value)
    sink_abundance = abundance_exchange(proportion_path, data_path, sources_num)
    noise_sink_data, sink_abundance, source_index, sources_num = data_clustering(sink_number, noise_sink_data, sink_abundance, sources_num, jsd_value)
    print('jensen shannon divergence between the original source data per source', source_jsd)
    print('jensen shannon between source data and noise data is', noise_source_jsd)
    print('classes generated by clustering is', source_index)
else:
    project_data, feast_data, noise_sink_data, source_jsd, sources_num, noise_source_jsd = data_output(data_path,
                                                                                                       sink_number,
                                                                                                       proportion_path,
                                                                                                       value)
    sink_abundance = abundance_exchange(proportion_path, data_path, sources_num)
    print('jensen shannon divergence between the original source data per source', source_jsd)
    print('jensen shannon between source data and noise data is', noise_source_jsd)


worksheet_project_data = data_to_excel(workbook_project_data, "project_data", 0, project_data)
workbook_project_data.save('./project_data.xlsx')
worksheet_feast_data = data_to_excel(workbook_feast_data, "feast_data", 0, feast_data)
workbook_feast_data.save('./feast_data.xlsx')
worksheet_noise_data = data_to_excel(workbook_noise_data, "noise_data", 0, noise_sink_data)
workbook_noise_data.save('./noise_data.xlsx')
worksheet_sink_abundance = data_to_excel(workbook_sink_abundance, 'sink_abundance', 0, sink_abundance)
workbook_sink_abundance.save('./sink_abundance.xlsx')


estimated_proportions = []  # Output the final estimated proportion
loss_sum = []  # Iterated value of the loss function
loss_final = []  # Final value of the loss function
diff_xwh_sum = []  # Difference value between X and WH
norm_xwh_sum = []  # Norm value between X and WH
diff_unknown_sum = []  # Difference value between UNKNOWN and original contribution
jsd_unknown_sum = []  # JSD between UNKNOWN and original contribution
curve_jsd = []  # Iterated curve of JSD
curve_xwh = []  # Difference curve between X and WH
curve_wo = []  # Difference curve between W and original contribution
curve_un = []  # Difference curve between UNKNOWN and original contribution


# data preprocessing and training
start_time = datetime.datetime.now()
for i in range(sink_number):
    remove_data, noise_data, sink_data, unknown_init, original_data = data_preprocess(i, noise_sink_data, sources_num)
    h_plus, w_plus, w, h, m, loss, loss_value, diffs_xwh, jsd, diffs_wo, diffs_un = run_admm(iteration,
                                                                                             sink_abundance[i, :],
                                                                                             noise_data,
                                                                                             sink_data,
                                                                                             original_data,
                                                                                             unknown_init,
                                                                                             rho,
                                                                                             threshold)
    print('The sink sample is run through', m, 'iterations to be completed')
    estimated_proportions.append(h_plus)
    loss_sum.append(loss)
    loss_final.append(loss_value)
    curve_jsd.append(jsd)
    curve_xwh.append(diffs_xwh)
    curve_wo.append(diffs_wo)
    curve_un.append(diffs_un)

    diff_xwh, norm_xwh, diff_unknown, jsd_unknown = diff_admm(sink_data, original_data, w_plus, w, h)
    diff_xwh_sum.append(diff_xwh)
    norm_xwh_sum.append(norm_xwh)
    diff_unknown_sum.append(diff_unknown)
    jsd_unknown_sum.append(jsd_unknown)

    worksheet_remove_data = data_to_excel(workbook_remove_data, "remove_data", i, remove_data)
    workbook_remove_data.save('./remove_data.xlsx')
    worksheet_w = data_to_excel(workbook_w, 'w', i, w)
    workbook_w.save('./w.xlsx')
    worksheet_h = data_to_excel(workbook_h, 'h', i, h)
    workbook_h.save('./h.xlsx')
    worksheet_w_plus = data_to_excel(workbook_w_plus, "w_plus", i, w_plus)
    workbook_w_plus.save('./w_plus.xlsx')


end_time = datetime.datetime.now()
var = (end_time - start_time).seconds
print('#### Runs for', var, 'seconds in total ####')


estimated_proportions = np.array(estimated_proportions)
jensen_shannon_mean, jensen_shannon, difference = perf_admm(estimated_proportions, sink_abundance)
worksheet_abundance = data_to_excel(workbook_abundance, "estimated_abundance", 0, estimated_proportions)
workbook_abundance.save('./estimated_abundance.xlsx')
print('jensen shannon divergence between estimated_proportions and sink_abundances is', jensen_shannon, jensen_shannon_mean)
print('difference between estimated_proportions and sink_abundances & average is', difference, np.mean(difference))
print('difference between x and np.dot(w * h) is', diff_xwh_sum)
print('matrix norm between x and np.dot(w * h) is', norm_xwh_sum)
print('difference between unknown and original contribution is', diff_unknown_sum)
print('jensen shannon divergence between unknown and original contribution is', jsd_unknown_sum)
print('threshold of the loss function is', loss_final)


perf_nmf = [jensen_shannon, difference, diff_xwh_sum, norm_xwh_sum, diff_unknown_sum, jsd_unknown_sum, loss_final]
perf_nmf = np.array(perf_nmf)
perf_nmf = pd.DataFrame(perf_nmf)
perf_nmf.to_excel('./perf_nmf.xlsx', index=False, header=False)


# plot_loss_function(loss_sum[0], loss_sum[1], loss_sum[2], loss_sum[3], './loss.jpg')
# plot_jsd_iteration(curve_jsd[0], curve_jsd[1], curve_jsd[2], curve_jsd[3], './jsd.jpg')
# plot_diff_xwh_iteration(curve_xwh[0], curve_xwh[1], curve_xwh[2], curve_xwh[3], './diff_xwh.jpg')
# plot_w_origin_iteration(curve_wo[0], curve_wo[1], curve_wo[2], curve_wo[3], './diff_wo.jpg')
# plot_unknown_iteration(curve_un[0], curve_un[1], curve_un[2], curve_un[3], './diff_unknown.jpg')





